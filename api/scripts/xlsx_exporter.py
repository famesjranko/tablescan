"""
xlsx_exporter.py
    Written for: TableScan 2.0 (US-018)
    Date: March 2026

Logic:
    Exports DataFrames to XLSX format with:
    - Merged cell support when structure_json contains span information
    - Distinct header row formatting
    - Proper cell alignment and borders

Args:
    dataframe: pandas DataFrame to export
    output_path: Path for the output XLSX file
    structure_json: Optional dict with cell span information
    header_rows: Number of header rows to format distinctly (default: 1)
"""

from typing import Optional
from pathlib import Path
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def export_to_xlsx(
    dataframe: pd.DataFrame,
    output_path: str,
    structure_json: Optional[dict] = None,
    header_rows: int = 1
) -> str:
    """
    Export a DataFrame to XLSX with cell spans and header formatting.

    Args:
        dataframe: pandas DataFrame to export
        output_path: Path for the output XLSX file
        structure_json: Optional dict with cell span information from extraction
        header_rows: Number of header rows to format distinctly

    Returns:
        str: Path to the created XLSX file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Table"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Determine number of header rows from structure_json if available
    if structure_json and "header_rows" in structure_json:
        header_rows = structure_json["header_rows"]

    # Write DataFrame to worksheet
    for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=False), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border

            # Apply header formatting for header rows
            if r_idx <= header_rows:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                cell.alignment = data_alignment

    # Auto-adjust column widths (with reasonable limits)
    # Must be done BEFORE applying cell spans to avoid MergedCell issues
    for col_idx in range(1, len(dataframe.columns) + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row_idx in range(1, len(dataframe) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with min/max constraints
        adjusted_width = min(max(max_length + 2, 8), 50)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Apply merged cells if structure_json contains span information
    if structure_json and "cells" in structure_json:
        _apply_cell_spans(ws, structure_json["cells"], header_rows, header_font, header_fill, header_alignment)

    # Apply hierarchical header merges if header_spans available
    if structure_json and "header_spans" in structure_json:
        _apply_header_spans(ws, structure_json["header_spans"], header_font, header_fill, header_alignment)

    # Save workbook
    wb.save(output_path)

    return output_path


def _apply_header_spans(
    ws,
    header_spans: list,
    header_font: Font,
    header_fill: PatternFill,
    header_alignment: Alignment
) -> None:
    """
    Apply merged cells for hierarchical headers based on header_spans.

    This creates merged cells for headers that span multiple columns (colspan)
    or multiple rows (rowspan), preserving the visual hierarchy from the PDF.

    Args:
        ws: openpyxl worksheet
        header_spans: List of span dicts with row, col, value, colspan, rowspan
        header_font: Font style for headers
        header_fill: Fill style for headers
        header_alignment: Alignment for headers
    """
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for span in header_spans:
        colspan = span.get("colspan", 1)
        rowspan = span.get("rowspan", 1)

        if colspan > 1 or rowspan > 1:
            # Convert 0-indexed to 1-indexed for openpyxl
            start_row = span["row"] + 1
            start_col = span["col"] + 1
            end_row = start_row + rowspan - 1
            end_col = start_col + colspan - 1

            # Merge cells
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col
            )

            # Re-apply formatting to merged cell
            merged_cell = ws.cell(row=start_row, column=start_col)
            merged_cell.value = span.get("value", "")
            merged_cell.font = header_font
            merged_cell.fill = header_fill
            merged_cell.alignment = header_alignment
            merged_cell.border = thin_border


def _apply_cell_spans(
    ws,
    cells: list,
    header_rows: int,
    header_font: Font,
    header_fill: PatternFill,
    header_alignment: Alignment
) -> None:
    """
    Apply cell merges based on span information from structure_json.

    Args:
        ws: openpyxl worksheet
        cells: List of cell dictionaries with row, col, value, rowspan, colspan
        header_rows: Number of header rows (for re-applying formatting after merge)
        header_font: Font style for headers
        header_fill: Fill style for headers
        header_alignment: Alignment for headers
    """
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for cell_info in cells:
        row = cell_info.get("row", 0)
        col = cell_info.get("col", 0)
        rowspan = cell_info.get("rowspan", 1)
        colspan = cell_info.get("colspan", 1)

        # Only process cells with spans > 1
        if rowspan > 1 or colspan > 1:
            # Convert 0-indexed to 1-indexed for openpyxl
            start_row = row + 1
            start_col = col + 1
            end_row = start_row + rowspan - 1
            end_col = start_col + colspan - 1

            # Merge cells
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=end_row,
                end_column=end_col
            )

            # Re-apply formatting to merged cell (merge resets some styles)
            merged_cell = ws.cell(row=start_row, column=start_col)
            merged_cell.border = thin_border
            merged_cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            # Re-apply header formatting if within header rows
            if start_row <= header_rows:
                merged_cell.font = header_font
                merged_cell.fill = header_fill


def build_xlsx_structure_from_extraction(
    dataframe: pd.DataFrame,
    extraction_metadata: Optional[dict] = None
) -> dict:
    """
    Build structure_json format suitable for XLSX export from extraction results.

    This helper can be used when extraction metadata contains cell-level span info
    from extractors like Camelot or img2table.

    Args:
        dataframe: The extracted DataFrame
        extraction_metadata: Optional metadata from the extractor

    Returns:
        dict with cells array and dimensions suitable for XLSX export
    """
    if dataframe is None or dataframe.empty:
        return None

    rows, cols = dataframe.shape
    cells = []

    for row_idx in range(rows):
        for col_idx in range(cols):
            cell_value = dataframe.iloc[row_idx, col_idx]
            cell_data = {
                "row": row_idx,
                "col": col_idx,
                "value": str(cell_value) if cell_value is not None else "",
            }

            # Check extraction metadata for span information
            if extraction_metadata:
                cells_meta = extraction_metadata.get("cells", [])
                for cm in cells_meta:
                    if cm.get("row") == row_idx and cm.get("col") == col_idx:
                        if cm.get("rowspan", 1) > 1:
                            cell_data["rowspan"] = cm["rowspan"]
                        if cm.get("colspan", 1) > 1:
                            cell_data["colspan"] = cm["colspan"]
                        break

            cells.append(cell_data)

    structure = {
        "rows": rows,
        "cols": cols,
        "cells": cells,
    }

    # Add header information if available
    if extraction_metadata and extraction_metadata.get("header_rows") is not None:
        structure["header_rows"] = extraction_metadata["header_rows"]

    return structure
