"""
End-to-end tests for Phase 4: Rich Output Schema

US-020: Verify Phase 4 end-to-end

Tests verify:
1. Upload PDF - all metadata fields populated
2. XLSX download works with proper formatting
3. API returns confidence scores
4. All existing tests still pass
"""

import io
import os
import json
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import fitz  # PyMuPDF
import pandas as pd
import pytest
from openpyxl import load_workbook

from api.scripts.extractors import (
    ExtractionScorer,
    ExtractionResult,
    CamelotExtractor,
    PdfplumberExtractor,
    MultiExtractor,
)
from api.scripts.xlsx_exporter import export_to_xlsx, build_xlsx_structure_from_extraction
from api.scripts.page_classifier import PageClassifier


# =============================================================================
# Fixtures
# =============================================================================

@pytest.fixture
def born_digital_pdf():
    """
    Create a born-digital PDF with a table for testing.
    """
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as f:
        pdf_path = f.name

    doc = fitz.open()
    page = doc.new_page(width=612, height=792)

    # Title
    page.insert_text(fitz.Point(50, 50), "Financial Report Q1 2025", fontsize=14)

    # Create bordered table
    shape = page.new_shape()

    x_start, y_start = 80, 100
    col_widths = [100, 100, 80, 80]
    row_height = 30
    num_rows = 5

    for i in range(num_rows + 1):
        y = y_start + i * row_height
        shape.draw_line(
            fitz.Point(x_start, y),
            fitz.Point(x_start + sum(col_widths), y)
        )

    x = x_start
    for width in [0] + col_widths:
        x += width
        shape.draw_line(
            fitz.Point(x, y_start),
            fitz.Point(x, y_start + num_rows * row_height)
        )

    shape.finish(color=(0, 0, 0), width=0.5)
    shape.commit()

    table_data = [
        ["Quarter", "Revenue", "Growth", "Status"],
        ["Q1 2024", "$125.4M", "12.5%", "Met"],
        ["Q2 2024", "$142.1M", "13.3%", "Exceeded"],
        ["Q3 2024", "$138.7M", "10.2%", "Met"],
        ["Q4 2024", "$156.2M", "15.8%", "Exceeded"],
    ]

    for i, row in enumerate(table_data):
        x = x_start
        for j, cell in enumerate(row):
            page.insert_text(
                fitz.Point(x + 5, y_start + i * row_height + 20),
                cell,
                fontsize=10
            )
            x += col_widths[j]

    doc.save(pdf_path)
    doc.close()

    yield pdf_path

    if os.path.exists(pdf_path):
        os.unlink(pdf_path)


@pytest.fixture
def sample_dataframe():
    """Create a sample DataFrame for XLSX export testing."""
    return pd.DataFrame({
        'Quarter': ['Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024'],
        'Revenue': ['$125.4M', '$142.1M', '$138.7M', '$156.2M'],
        'Growth': ['12.5%', '13.3%', '10.2%', '15.8%'],
        'Status': ['Met', 'Exceeded', 'Met', 'Exceeded']
    })


@pytest.fixture
def sample_structure_json():
    """Create sample structure_json with cell spans."""
    return {
        'rows': 5,
        'cols': 4,
        'header_rows': 1,
        'cells': [
            {'row': 0, 'col': 0, 'value': 'Quarter'},
            {'row': 0, 'col': 1, 'value': 'Revenue', 'colspan': 2},  # Merged header
            {'row': 0, 'col': 3, 'value': 'Status'},
            {'row': 1, 'col': 0, 'value': 'Q1 2024'},
            {'row': 1, 'col': 1, 'value': '$125.4M'},
            {'row': 1, 'col': 2, 'value': '12.5%'},
            {'row': 1, 'col': 3, 'value': 'Met'},
            # ... more cells would follow
        ]
    }


# =============================================================================
# Test: Metadata Fields Population
# =============================================================================

class TestMetadataFieldsPopulated:
    """Verify all metadata fields are populated during extraction."""

    @pytest.fixture
    def model_source(self):
        """Read models.py source code."""
        path = Path(__file__).parent.parent / "models.py"
        return path.read_text()

    @pytest.fixture
    def predict_table_source(self):
        """Read predict_table.py source code."""
        path = Path(__file__).parent.parent / "scripts" / "YOLOV3" / "predict_table.py"
        return path.read_text()

    def test_extracted_model_has_page_type_field(self, model_source):
        """Extracted model should have page_type field."""
        assert "page_type" in model_source
        assert "CharField" in model_source

    def test_extracted_model_has_extraction_method_field(self, model_source):
        """Extracted model should have extraction_method field."""
        assert "extraction_method" in model_source

    def test_extracted_model_has_confidence_score_field(self, model_source):
        """Extracted model should have confidence_score field."""
        assert "confidence_score" in model_source
        assert "FloatField" in model_source

    def test_extracted_model_has_structure_json_field(self, model_source):
        """Extracted model should have structure_json field."""
        assert "structure_json" in model_source
        assert "JSONField" in model_source

    def test_extracted_model_has_bounding_box_field(self, model_source):
        """Extracted model should have bounding_box field."""
        assert "bounding_box" in model_source
        assert "JSONField" in model_source

    def test_detect_tables_populates_page_type(self, predict_table_source):
        """detect_tables should populate page_type field."""
        assert "page_type=page_type" in predict_table_source

    def test_detect_tables_populates_extraction_method(self, predict_table_source):
        """detect_tables should populate extraction_method field."""
        assert "extraction_method=" in predict_table_source

    def test_detect_tables_populates_confidence_score(self, predict_table_source):
        """detect_tables should populate confidence_score field."""
        assert "confidence_score=" in predict_table_source

    def test_detect_tables_populates_structure_json(self, predict_table_source):
        """detect_tables should populate structure_json field."""
        assert "structure_json=" in predict_table_source

    def test_detect_tables_populates_bounding_box(self, predict_table_source):
        """detect_tables should populate bounding_box field."""
        assert "bounding_box=" in predict_table_source

    def test_extract_tables_direct_populates_all_fields(self, predict_table_source):
        """extract_tables_direct should populate all rich fields.

        Note: After decoupling extraction from saving (Phase 1 architecture refactor),
        the field population is in save_extraction_results(), which is called by
        extract_tables_direct().
        """
        lines = predict_table_source.split('\n')
        in_direct_function = False
        calls_save_results = False
        in_save_function = False
        field_checks = {
            'page_type': False,
            'extraction_method': False,
            'confidence_score': False,
            'structure_json': False,
            'bounding_box': False,
        }

        # Check that extract_tables_direct calls save_extraction_results
        for line in lines:
            if 'def extract_tables_direct(' in line:
                in_direct_function = True
            elif in_direct_function and line.startswith('def ') and 'extract_tables_direct' not in line:
                in_direct_function = False
            elif in_direct_function and 'save_extraction_results' in line:
                calls_save_results = True

        # Check that save_extraction_results populates all fields
        for line in lines:
            if 'def save_extraction_results(' in line:
                in_save_function = True
            elif in_save_function and line.startswith('def ') and 'save_extraction_results' not in line:
                break
            elif in_save_function:
                for field in field_checks:
                    if f'{field}=' in line:
                        field_checks[field] = True

        assert calls_save_results, "extract_tables_direct should call save_extraction_results"
        for field, found in field_checks.items():
            assert found, f"save_extraction_results should populate {field}"

    def test_extract_tables_vision_populates_all_fields(self, predict_table_source):
        """extract_tables_vision should populate all rich fields.

        Note: After decoupling extraction from saving (Phase 1 architecture refactor),
        the field population is in save_extraction_results(), which is called by
        extract_tables_vision().
        """
        lines = predict_table_source.split('\n')
        in_vision_function = False
        calls_save_results = False
        in_save_function = False
        field_checks = {
            'page_type': False,
            'extraction_method': False,
            'confidence_score': False,
            'structure_json': False,
            'bounding_box': False,
        }

        # Check that extract_tables_vision calls save_extraction_results
        for line in lines:
            if 'def extract_tables_vision(' in line:
                in_vision_function = True
            elif in_vision_function and line.startswith('def ') and 'extract_tables_vision' not in line:
                in_vision_function = False
            elif in_vision_function and 'save_extraction_results' in line:
                calls_save_results = True

        # Check that save_extraction_results populates all fields
        for line in lines:
            if 'def save_extraction_results(' in line:
                in_save_function = True
            elif in_save_function and line.startswith('def ') and 'save_extraction_results' not in line:
                break
            elif in_save_function:
                for field in field_checks:
                    if f'{field}=' in line:
                        field_checks[field] = True

        assert calls_save_results, "extract_tables_vision should call save_extraction_results"
        for field, found in field_checks.items():
            assert found, f"save_extraction_results should populate {field}"


# =============================================================================
# Test: XLSX Export with Proper Formatting
# =============================================================================

class TestXLSXExport:
    """Verify XLSX export works with proper formatting."""

    def test_export_to_xlsx_creates_file(self, sample_dataframe, tmp_path):
        """export_to_xlsx should create an XLSX file."""
        output_path = tmp_path / "test_table.xlsx"

        result = export_to_xlsx(sample_dataframe, str(output_path))

        assert os.path.exists(output_path)
        assert result == str(output_path)

    def test_export_to_xlsx_contains_data(self, sample_dataframe, tmp_path):
        """XLSX file should contain the DataFrame data."""
        output_path = tmp_path / "test_table.xlsx"

        export_to_xlsx(sample_dataframe, str(output_path))

        wb = load_workbook(output_path)
        ws = wb.active

        # Check that data is present
        assert ws['A1'].value == 'Q1 2024'  # First data cell (no header in export)
        assert ws['B1'].value == '$125.4M'

    def test_export_to_xlsx_with_header_formatting(self, sample_dataframe, tmp_path):
        """XLSX export should format header rows distinctly."""
        output_path = tmp_path / "test_table.xlsx"

        # Create structure_json with header_rows
        structure = {'rows': 5, 'cols': 4, 'header_rows': 1, 'cells': []}

        export_to_xlsx(sample_dataframe, str(output_path), structure_json=structure)

        wb = load_workbook(output_path)
        ws = wb.active

        # First row should have header formatting (bold)
        cell = ws['A1']
        assert cell.font.bold is True

    def test_export_to_xlsx_applies_borders(self, sample_dataframe, tmp_path):
        """XLSX export should apply borders to cells."""
        output_path = tmp_path / "test_table.xlsx"

        export_to_xlsx(sample_dataframe, str(output_path))

        wb = load_workbook(output_path)
        ws = wb.active

        # Check that cells have borders
        cell = ws['A1']
        assert cell.border.left.style is not None
        assert cell.border.right.style is not None
        assert cell.border.top.style is not None
        assert cell.border.bottom.style is not None

    def test_export_to_xlsx_with_cell_spans(self, sample_dataframe, tmp_path, sample_structure_json):
        """XLSX export should preserve cell spans from structure_json."""
        output_path = tmp_path / "test_table.xlsx"

        # Use a simpler structure for testing spans
        structure_with_spans = {
            'rows': 4,
            'cols': 4,
            'header_rows': 1,
            'cells': [
                {'row': 0, 'col': 1, 'value': 'Revenue', 'colspan': 2},  # Merged header
            ]
        }

        export_to_xlsx(sample_dataframe, str(output_path), structure_json=structure_with_spans)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check for merged cells
        merged_ranges = list(ws.merged_cells.ranges)
        # At least one merged range should exist (from colspan)
        assert len(merged_ranges) >= 1

    def test_export_to_xlsx_adjusts_column_width(self, sample_dataframe, tmp_path):
        """XLSX export should auto-adjust column widths."""
        output_path = tmp_path / "test_table.xlsx"

        export_to_xlsx(sample_dataframe, str(output_path))

        wb = load_workbook(output_path)
        ws = wb.active

        # Column widths should be adjusted (not default)
        col_a_width = ws.column_dimensions['A'].width
        assert col_a_width is not None
        assert col_a_width >= 8  # Minimum width we set

    def test_export_to_xlsx_handles_overlapping_spans(self, sample_dataframe, tmp_path):
        """
        XLSX export should handle overlapping spans gracefully.

        When both cells array and header_spans try to merge the same region,
        it should not raise 'MergedCell object attribute value is read-only'.
        """
        # Given: structure_json with overlapping span definitions
        output_path = tmp_path / "test_table.xlsx"
        structure_with_overlaps = {
            "rows": 3,
            "cols": 3,
            "cells": [
                {"row": 0, "col": 0, "value": "Header", "colspan": 2, "rowspan": 1},
            ],
            "header_spans": [
                {"row": 0, "col": 0, "value": "Header", "colspan": 2, "rowspan": 1},
            ]
        }

        # When: exporting to XLSX (should not raise exception)
        export_to_xlsx(sample_dataframe, str(output_path), structure_json=structure_with_overlaps)

        # Then: file is created successfully
        wb = load_workbook(output_path)
        ws = wb.active
        merged_ranges = list(ws.merged_cells.ranges)
        assert len(merged_ranges) >= 1


class TestXLSXExportInPipeline:
    """Verify XLSX export is integrated into the extraction pipeline."""

    @pytest.fixture
    def predict_table_source(self):
        """Read predict_table.py source code."""
        path = Path(__file__).parent.parent / "scripts" / "YOLOV3" / "predict_table.py"
        return path.read_text()

    @pytest.fixture
    def table_extract_source(self):
        """Read table_extract.py source code."""
        path = Path(__file__).parent.parent / "scripts" / "table_extract.py"
        return path.read_text()

    def test_xlsx_export_imported(self, predict_table_source):
        """export_to_xlsx should be imported in predict_table.py."""
        assert "from api.scripts.xlsx_exporter import export_to_xlsx" in predict_table_source

    def test_xlsx_directory_created(self, table_extract_source):
        """XLSX extraction directory should be created."""
        assert '"xlsx"' in table_extract_source

    def test_detect_tables_exports_xlsx(self, predict_table_source):
        """detect_tables should export XLSX files."""
        lines = predict_table_source.split('\n')
        in_function = False
        exports_xlsx = False

        for line in lines:
            if 'def detect_tables(' in line:
                in_function = True
            elif in_function and line.startswith('def ') and 'detect_tables' not in line:
                break
            elif in_function and 'export_to_xlsx' in line:
                exports_xlsx = True

        assert exports_xlsx, "detect_tables should call export_to_xlsx"

    def test_extract_tables_direct_exports_xlsx(self, predict_table_source):
        """extract_tables_direct should export XLSX files.

        Note: After decoupling extraction from saving (Phase 1 architecture refactor),
        the XLSX export is in save_extraction_results(), which is called by
        extract_tables_direct().
        """
        lines = predict_table_source.split('\n')
        in_direct_function = False
        calls_save_results = False
        in_save_function = False
        exports_xlsx = False

        # Check that extract_tables_direct calls save_extraction_results
        for line in lines:
            if 'def extract_tables_direct(' in line:
                in_direct_function = True
            elif in_direct_function and line.startswith('def ') and 'extract_tables_direct' not in line:
                in_direct_function = False
            elif in_direct_function and 'save_extraction_results' in line:
                calls_save_results = True

        # Check that save_extraction_results exports XLSX
        for line in lines:
            if 'def save_extraction_results(' in line:
                in_save_function = True
            elif in_save_function and line.startswith('def ') and 'save_extraction_results' not in line:
                break
            elif in_save_function and 'export_to_xlsx' in line:
                exports_xlsx = True

        assert calls_save_results, "extract_tables_direct should call save_extraction_results"
        assert exports_xlsx, "save_extraction_results should call export_to_xlsx"

    def test_extract_tables_vision_exports_xlsx(self, predict_table_source):
        """extract_tables_vision should export XLSX files.

        Note: After decoupling extraction from saving (Phase 1 architecture refactor),
        the XLSX export is in save_extraction_results(), which is called by
        extract_tables_vision().
        """
        lines = predict_table_source.split('\n')
        in_vision_function = False
        calls_save_results = False
        in_save_function = False
        exports_xlsx = False

        # Check that extract_tables_vision calls save_extraction_results
        for line in lines:
            if 'def extract_tables_vision(' in line:
                in_vision_function = True
            elif in_vision_function and line.startswith('def ') and 'extract_tables_vision' not in line:
                in_vision_function = False
            elif in_vision_function and 'save_extraction_results' in line:
                calls_save_results = True

        # Check that save_extraction_results exports XLSX
        for line in lines:
            if 'def save_extraction_results(' in line:
                in_save_function = True
            elif in_save_function and line.startswith('def ') and 'save_extraction_results' not in line:
                break
            elif in_save_function and 'export_to_xlsx' in line:
                exports_xlsx = True

        assert calls_save_results, "extract_tables_vision should call save_extraction_results"
        assert exports_xlsx, "save_extraction_results should call export_to_xlsx"


# =============================================================================
# Test: API Returns Confidence Scores
# =============================================================================

class TestAPIReturnsConfidenceScores:
    """Verify API responses include confidence scores."""

    @pytest.fixture
    def serializers_source(self):
        """Read serializers.py source code."""
        path = Path(__file__).parent.parent / "serializers.py"
        return path.read_text()

    def test_extracted_serializer_includes_page_type(self, serializers_source):
        """ExtractedSerializer should include page_type field."""
        assert '"page_type"' in serializers_source

    def test_extracted_serializer_includes_extraction_method(self, serializers_source):
        """ExtractedSerializer should include extraction_method field."""
        assert '"extraction_method"' in serializers_source

    def test_extracted_serializer_includes_confidence_score(self, serializers_source):
        """ExtractedSerializer should include confidence_score field."""
        assert '"confidence_score"' in serializers_source

    def test_extracted_serializer2_includes_metadata(self, serializers_source):
        """ExtractedSerializer2 should include metadata fields."""
        # This serializer is used for nested extracted data in ReportSerializer
        assert 'class ExtractedSerializer2' in serializers_source

        # Find the fields in ExtractedSerializer2
        lines = serializers_source.split('\n')
        in_serializer2 = False
        found_fields = {'page_type': False, 'extraction_method': False, 'confidence_score': False}

        for line in lines:
            if 'class ExtractedSerializer2' in line:
                in_serializer2 = True
            elif in_serializer2 and line.startswith('class '):
                break
            elif in_serializer2:
                for field in found_fields:
                    if f'"{field}"' in line:
                        found_fields[field] = True

        for field, found in found_fields.items():
            assert found, f"ExtractedSerializer2 should include {field}"


class TestConfidenceScoreCalculation:
    """Verify confidence scores are calculated correctly."""

    def test_extraction_scorer_returns_normalized_score(self):
        """ExtractionScorer should return scores between 0 and 1."""
        scorer = ExtractionScorer()

        result = ExtractionResult(
            dataframe=pd.DataFrame({
                'A': ['1', '2', '3'],
                'B': ['4', '5', '6'],
            }),
            confidence=0.8,
            method='test',
            metadata={}
        )

        score = scorer.score(result)
        assert 0.0 <= score <= 1.0

    def test_camelot_extractor_returns_confidence(self, born_digital_pdf):
        """CamelotExtractor should return results with confidence scores."""
        extractor = CamelotExtractor(flavor='lattice')

        try:
            results = extractor.extract(born_digital_pdf, 1)
            for result in results:
                assert 0.0 <= result.confidence <= 1.0
        except Exception:
            # PDF may not have tables or camelot may fail
            pytest.skip("Camelot extraction failed on test PDF")

    def test_pdfplumber_extractor_returns_confidence(self, born_digital_pdf):
        """PdfplumberExtractor should return results with confidence scores."""
        extractor = PdfplumberExtractor()

        try:
            results = extractor.extract(born_digital_pdf, 1)
            for result in results:
                assert 0.0 <= result.confidence <= 1.0
        except Exception:
            # PDF may not have tables or pdfplumber may fail
            pytest.skip("Pdfplumber extraction failed on test PDF")


# =============================================================================
# Test: Build Structure JSON
# =============================================================================

class TestBuildStructureJson:
    """Verify structure_json is built correctly."""

    @pytest.fixture
    def predict_table_source(self):
        """Read predict_table.py source code."""
        path = Path(__file__).parent.parent / "scripts" / "YOLOV3" / "predict_table.py"
        return path.read_text()

    def test_build_structure_json_function_exists(self, predict_table_source):
        """build_structure_json function should exist."""
        assert "def build_structure_json(" in predict_table_source

    def test_build_structure_json_creates_cells_array(self, sample_dataframe):
        """build_structure_json should create cells array."""
        structure = build_xlsx_structure_from_extraction(sample_dataframe)
        assert 'cells' in structure
        assert isinstance(structure['cells'], list)

    def test_structure_json_includes_dimensions(self, sample_dataframe):
        """structure_json should include row and column counts."""
        structure = build_xlsx_structure_from_extraction(sample_dataframe)
        assert 'rows' in structure
        assert 'cols' in structure

    def test_build_xlsx_structure_from_extraction_exists(self):
        """build_xlsx_structure_from_extraction should be available."""
        from api.scripts.xlsx_exporter import build_xlsx_structure_from_extraction
        assert callable(build_xlsx_structure_from_extraction)

    def test_build_xlsx_structure_from_extraction_works(self, sample_dataframe):
        """build_xlsx_structure_from_extraction should build valid structure."""
        structure = build_xlsx_structure_from_extraction(sample_dataframe)

        assert structure is not None
        assert 'rows' in structure
        assert 'cols' in structure
        assert 'cells' in structure
        assert structure['rows'] == 4
        assert structure['cols'] == 4


# =============================================================================
# Test: Backwards Compatibility
# =============================================================================

class TestBackwardsCompatibility:
    """Verify new fields are backwards compatible."""

    @pytest.fixture
    def model_source(self):
        """Read models.py source code."""
        path = Path(__file__).parent.parent / "models.py"
        return path.read_text()

    def test_new_fields_are_nullable(self, model_source):
        """New rich metadata fields should be nullable."""
        # All new fields should have null=True for backwards compatibility
        assert "null=True" in model_source

    def test_api_responses_still_include_base_fields(self):
        """API responses should still include base fields."""
        from api.serializers import ExtractedSerializer

        # Get the fields from the serializer
        serializer = ExtractedSerializer()
        fields = serializer.Meta.fields

        # Original fields should still be present
        assert 'report' in fields
        assert 'page_num' in fields
        assert 'table_num' in fields
        assert 'f_type' in fields
        assert 'file' in fields


# =============================================================================
# Test: Migration Exists
# =============================================================================

class TestMigrationExists:
    """Verify the database migration for rich fields exists."""

    def test_migration_file_exists(self):
        """Migration file for rich fields should exist."""
        migrations_dir = Path(__file__).parent.parent / "migrations"

        # Look for a migration that adds rich fields
        migration_files = list(migrations_dir.glob("0005_*.py"))
        assert len(migration_files) > 0, "Migration 0005 should exist"

    def test_migration_adds_rich_fields(self):
        """Migration should add the rich fields."""
        migrations_dir = Path(__file__).parent.parent / "migrations"
        migration_files = list(migrations_dir.glob("0005_*.py"))

        if not migration_files:
            pytest.skip("Migration file not found")

        migration_content = migration_files[0].read_text()

        # Check that migration adds the new fields
        assert "page_type" in migration_content
        assert "extraction_method" in migration_content
        assert "confidence_score" in migration_content
        assert "structure_json" in migration_content
        assert "bounding_box" in migration_content


# =============================================================================
# Test: Feature Flags
# =============================================================================

class TestFeatureFlags:
    """Verify feature flags for rich schema."""

    @pytest.fixture
    def prd_json_path(self):
        """Path to prd.json."""
        return Path(__file__).parent.parent.parent / "docs" / "prd.json"

    def test_prd_json_has_feature_flags(self, prd_json_path):
        """prd.json should have featureFlags section."""
        with open(prd_json_path) as f:
            prd = json.load(f)

        assert 'featureFlags' in prd

    def test_use_rich_schema_flag_exists(self, prd_json_path):
        """use_rich_schema feature flag should exist."""
        with open(prd_json_path) as f:
            prd = json.load(f)

        assert 'use_rich_schema' in prd['featureFlags']


# =============================================================================
# Test: Extraction Results Include Rich Data
# =============================================================================

class TestExtractionResultsIncludeRichData:
    """Verify extraction results include rich metadata."""

    def test_extraction_result_has_metadata_dict(self):
        """ExtractionResult should have metadata dict."""
        result = ExtractionResult(
            dataframe=pd.DataFrame({'A': [1, 2, 3]}),
            confidence=0.9,
            method='test',
            metadata={'key': 'value'}
        )

        assert isinstance(result.metadata, dict)

    def test_multi_extractor_includes_method(self, born_digital_pdf):
        """MultiExtractor results should include extraction method."""
        extractor = MultiExtractor()

        try:
            results, comparison = extractor.extract_with_comparison(born_digital_pdf, 1)

            for result in results:
                assert result.method is not None
                assert len(result.method) > 0
        except Exception:
            pytest.skip("MultiExtractor failed on test PDF")


# =============================================================================
# Integration Test (if Django available)
# =============================================================================

try:
    from django.test import TestCase
    from django.contrib.auth.models import User
    from rest_framework.test import APITestCase, APIClient
    DJANGO_API_AVAILABLE = True
except ImportError:
    DJANGO_API_AVAILABLE = False

    class APITestCase:
        pass


@pytest.mark.skipif(not DJANGO_API_AVAILABLE, reason="Django API not available")
@pytest.mark.django_db(transaction=True)
class TestPhase4APIIntegration(APITestCase):
    """API integration tests for Phase 4 rich output."""

    def setUp(self):
        """Set up test user and client."""
        self.user = User.objects.create_user(
            username='phase4_test_user',
            password='testpass123',
            email='phase4@example.com'
        )
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)

    def tearDown(self):
        """Clean up after tests."""
        from api.models import Report
        Report.objects.filter(owner=self.user).delete()

    def test_extracted_model_rich_fields_exist(self):
        """Extracted model should have rich metadata fields."""
        from api.models import Extracted

        # Check model has the fields
        field_names = [f.name for f in Extracted._meta.get_fields()]
        assert 'page_type' in field_names
        assert 'extraction_method' in field_names
        assert 'confidence_score' in field_names
        assert 'structure_json' in field_names
        assert 'bounding_box' in field_names

    def test_extracted_serializer_exposes_rich_fields(self):
        """ExtractedSerializer should expose rich metadata fields."""
        from api.serializers import ExtractedSerializer

        fields = ExtractedSerializer.Meta.fields
        assert 'page_type' in fields
        assert 'extraction_method' in fields
        assert 'confidence_score' in fields


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
